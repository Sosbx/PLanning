# © 2024 HILAL Arkane. Tous droits réservés.
# gui/planning_management.py

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QTextEdit, QFileDialog, QMessageBox, QListWidget)
from PyQt6.QtCore import Qt, QDate, QTimer
import os
import json
import csv
import openpyxl
from datetime import datetime, timedelta, date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from calendar import monthrange, day_abbr
from openpyxl.utils import get_column_letter
from core.Constantes.models import Planning, DayPlanning, TimeSlot
from workalendar.europe import France

ORDERED_POST_TYPES = [
    "MLD1", "ACD1", "MCD1", "ALD1", "MLD2", "ACD2", "MCD2", "ALD2", "MLD3", "ACD3", "MCD3", "ALD3",
    "MLD4", "ACD4", "MCD4", "ALD4", "MLD5", "ACD5", "MCD5", "ALD5", "MMD1", "MMD2", "AMD1", "NLD1",
    "NLD2", "NMD1", "NMD2", "NMD3", "NAD1", "NAD2", "NAD3", "NZD1", "NCD1", "NCD2", "NCD3", "SMD1",
    "SAD1", "SSD1", "RMD1", "RAD1", "RSD1", "HMD1", "HAD1", "HSD1", "CMD1", "CAD1", "CSD1", "CMD2",
    "CAD2", "CSD2", "CMD3", "CAD3", "CMD4", "CAD4", "CTD1"
]
POST_TYPE_MAPPING = {
    "ML": "MLD", "AC": "ACD", "MC": "MCD", "AL": "ALD", "MM": "MMD", "AM": "AMD",
    "NL": "NLD", "NM": "NMD", "NA": "NAD", "NC": "NCD", "SM": "SMD", "SA": "SAD",
    "SS": "SSD", "RM": "RMD", "RA": "RAD", "RS": "RSD", "HM": "HMD", "HA": "HAD",
    "HS": "HSD", "CM": "CMD", "CA": "CAD", "CS": "CSD", "CT": "CTD"
}
class PlanningManagementWidget(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.cal = France()  # Initialisation du calendrier ici
        
        # Configuration de la sauvegarde automatique
        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.timeout.connect(self.auto_save_planning)
        self.auto_save_filename = os.path.join(os.getcwd(), "auto_save_planning.json")
        
        # Connecter l'arrêt du timer à la fermeture de la fenêtre principale
        self.main_window.destroyed.connect(self.stop_auto_save)
        
        self.init_ui()

    def stop_auto_save(self):
        """Arrête le timer de sauvegarde automatique"""
        if self.auto_save_timer.isActive():
            self.auto_save_timer.stop()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Boutons avec style professionnel
        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)

        self.save_button = QPushButton("Sauvegarder le planning")
        self.load_button = QPushButton("Charger un planning")
        self.export_csv_button = QPushButton("Exporter en CSV")
        self.export_excel_button = QPushButton("Exporter en Excel")
        
        # Appliquer le style des boutons d'action
        for button in [self.save_button, self.load_button, self.export_csv_button, self.export_excel_button]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #2c5282;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 10pt;
                }
                QPushButton:hover {
                    background-color: #1a365d;
                }
                QPushButton:pressed {
                    background-color: #2a4365;
                }
            """)
            button.setMinimumWidth(150)
        
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.load_button)
        button_layout.addWidget(self.export_csv_button)
        button_layout.addWidget(self.export_excel_button)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)

        # Liste des plannings avec style professionnel
        self.planning_list = QListWidget()
        self.planning_list.setStyleSheet("""
            QListWidget {
                background-color: #f8fafc;
                border: 1px solid #c0d0e0;
                border-radius: 4px;
                padding: 5px;
                font-size: 10pt;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e2e8f0;
            }
            QListWidget::item:selected {
                background-color: #e8f0f8;
                color: #2c5282;
            }
            QListWidget::item:hover {
                background-color: #f0f5fa;
            }
        """)
        layout.addWidget(self.planning_list)

        # Connecter les boutons aux fonctions
        self.save_button.clicked.connect(self.save_planning)
        self.load_button.clicked.connect(self.load_planning)
        self.export_csv_button.clicked.connect(self.export_to_csv)
        self.export_excel_button.clicked.connect(self.export_to_excel)

        # Mettre à jour la liste des plannings
        self.update_planning_list()

    def start_auto_save(self):
        """Démarre la sauvegarde automatique"""
        if not self.auto_save_timer.isActive():
            self.auto_save_planning()  # Sauvegarde initiale
            self.auto_save_timer.start(5 * 60 * 1000)  # 5 minutes en millisecondes

    def auto_save_planning(self):
        """Méthode pour la sauvegarde automatique du planning"""
        if not self.main_window.planning_tab.planning:
            return
        
        print(f"Sauvegarde automatique en cours vers {self.auto_save_filename}")

        planning = self.main_window.planning_tab.planning
        
        # Fonction pour convertir les clés en format sérialisable
        def convert_dict_keys_to_iso(d):
            if isinstance(d, dict):
                return {
                    (k.isoformat() if isinstance(k, (date, datetime))
                     else '|'.join(map(str, k)) if isinstance(k, tuple)
                     else k): convert_dict_keys_to_iso(v)
                    for k, v in d.items()
                }
            elif isinstance(d, list):
                return [convert_dict_keys_to_iso(item) for item in d]
            elif isinstance(d, tuple):
                return '|'.join(map(str, d))
            return d

        data = {
            "start_date": planning.start_date,
            "end_date": planning.end_date,
            "pre_analysis_results": convert_dict_keys_to_iso(planning.pre_analysis_results),
            "weekend_validated": planning.weekend_validated,
            "pre_attributions": self.main_window.planning_tab.pre_attribution_tab.pre_attributions,
            "days": [
                {
                    "date": day.date,
                    "slots": [
                        {
                            "start_time": slot.start_time,
                            "end_time": slot.end_time,
                            "site": slot.site,
                            "slot_type": slot.slot_type,
                            "abbreviation": slot.abbreviation,
                            "assignee": slot.assignee
                        } for slot in day.slots
                    ]
                } for day in planning.days
            ]
        }

        class DateTimeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                elif isinstance(obj, tuple):
                    return '|'.join(map(str, obj))
                return super().default(obj)

            def encode(self, obj):
                if isinstance(obj, dict):
                    # Convert dictionary with potential tuple keys
                    new_dict = {}
                    for k, v in obj.items():
                        if isinstance(k, tuple):
                            k = '|'.join(map(str, k))
                        new_dict[k] = v
                    return super().encode(new_dict)
                return super().encode(obj)

        # First convert all nested structures
        converted_data = convert_dict_keys_to_iso(data)
        
        # Then save with the enhanced encoder
        with open(self.auto_save_filename, 'w') as f:
            json.dump(converted_data, f, cls=DateTimeEncoder)

    def save_planning(self):
        if not self.main_window.planning_tab.planning:
            QMessageBox.warning(self, "Erreur", "Aucun planning à sauvegarder.")
            return
            
        # Démarrer la sauvegarde automatique
        self.start_auto_save()

        planning = self.main_window.planning_tab.planning
        
        # Si le planning n'a pas encore de fichier associé, demander où sauvegarder
        if not planning.filename:
            default_name = f"P {planning.start_date.strftime('%b')} - {planning.end_date.strftime('%b')} - {planning.end_date.year}.json"
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Sauvegarder le planning",
                default_name,
                "Fichiers JSON (*.json)"
            )
            if not filename:
                return
            planning.filename = filename
        
        # Fonction pour convertir les clés en format sérialisable
        def convert_dict_keys_to_iso(d):
            if isinstance(d, dict):
                return {
                    (k.isoformat() if isinstance(k, (date, datetime))
                     else '|'.join(map(str, k)) if isinstance(k, tuple)
                     else k): convert_dict_keys_to_iso(v)
                    for k, v in d.items()
                }
            elif isinstance(d, list):
                return [convert_dict_keys_to_iso(item) for item in d]
            elif isinstance(d, tuple):
                return '|'.join(map(str, d))
            return d

        data = {
            "start_date": planning.start_date,
            "end_date": planning.end_date,
            "pre_analysis_results": convert_dict_keys_to_iso(planning.pre_analysis_results),
            "weekend_validated": planning.weekend_validated,  # Save weekend validation state
            "pre_attributions": self.main_window.planning_tab.pre_attribution_tab.pre_attributions,
            "days": [
                {
                    "date": day.date,
                    "slots": [
                        {
                            "start_time": slot.start_time,
                            "end_time": slot.end_time,
                            "site": slot.site,
                            "slot_type": slot.slot_type,
                            "abbreviation": slot.abbreviation,
                            "assignee": slot.assignee
                        } for slot in day.slots
                    ]
                } for day in planning.days
            ]
        }

        class DateTimeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                elif isinstance(obj, tuple):
                    return '|'.join(map(str, obj))
                return super().default(obj)

            def encode(self, obj):
                if isinstance(obj, dict):
                    # Convert dictionary with potential tuple keys
                    new_dict = {}
                    for k, v in obj.items():
                        if isinstance(k, tuple):
                            k = '|'.join(map(str, k))
                        new_dict[k] = v
                    return super().encode(new_dict)
                return super().encode(obj)

        # First convert all nested structures
        converted_data = convert_dict_keys_to_iso(data)
        
        # Then save with the enhanced encoder
        with open(planning.filename, 'w') as f:
            json.dump(converted_data, f, cls=DateTimeEncoder)

        self.update_planning_list()
        QMessageBox.information(self, "Succès", f"Planning sauvegardé sous {planning.filename}")

    def load_planning(self):
        filename, _ = QFileDialog.getOpenFileName(self, "Charger un planning", "", "Fichiers JSON (*.json)")
        if not filename:
            return

        with open(filename, 'r') as f:
            data = json.load(f)

        planning = Planning(
            start_date=datetime.fromisoformat(data['start_date']).date(),
            end_date=datetime.fromisoformat(data['end_date']).date(),
            filename=filename
        )
        planning.weekend_validated = data.get('weekend_validated', False)  # Load weekend validation state
        
        # Restaurer les pré-attributions
        if 'pre_attributions' in data:
            self.main_window.planning_tab.pre_attribution_tab.pre_attributions = data['pre_attributions']

        for day_data in data['days']:
            day_date = datetime.fromisoformat(day_data['date']).date()
            day = DayPlanning(date=day_date)
            day.planning = planning  # Lier le jour au planning
            # Marquer les weekends et jours fériés
            day.is_weekend = day_date.weekday() >= 5
            day.is_holiday_or_bridge = self.cal.is_holiday(day_date) or self.is_bridge_day(day_date)
            for slot_data in day_data['slots']:
                slot = TimeSlot(
                    start_time=datetime.fromisoformat(slot_data['start_time']),
                    end_time=datetime.fromisoformat(slot_data['end_time']),
                    site=slot_data['site'],
                    slot_type=slot_data['slot_type'],
                    abbreviation=slot_data['abbreviation'],
                    assignee=slot_data['assignee']
                )
                day.slots.append(slot)
            planning.days.append(day)

        # Fonction pour convertir les clés et valeurs en dates
        def convert_dict_from_iso(d):
            if not isinstance(d, dict):
                return d
            new_dict = {}
            for k, v in d.items():
                try:
                    # Try to convert to date first
                    key = datetime.fromisoformat(k).date()
                except (ValueError, TypeError):
                    # If not a date, check if it's a serialized tuple
                    if isinstance(k, str) and '|' in k:
                        try:
                            parts = k.split('|')
                            # Attempt to convert parts to int if possible
                            key = tuple(int(p) if p.isdigit() else p for p in parts)
                        except (ValueError, TypeError):
                            key = k
                    else:
                        key = k
                
                if isinstance(v, dict):
                    value = convert_dict_from_iso(v)
                elif isinstance(v, list):
                    value = [convert_dict_from_iso(item) for item in v]
                elif isinstance(v, str):
                    try:
                        value = datetime.fromisoformat(v).date()
                    except (ValueError, TypeError):
                        value = v
                else:
                    value = v
                new_dict[key] = value
            return new_dict

        if 'pre_analysis_results' in data:
            planning.pre_analysis_results = convert_dict_from_iso(data['pre_analysis_results'])
        else:
            planning.pre_analysis_results = {}

        self.main_window.planning_tab.planning = planning
        
        # Démarrer la sauvegarde automatique
        self.start_auto_save()
        
        # Mettre à jour les dates dans PlanningViewWidget
        self.main_window.planning_tab.start_date.setDate(QDate(planning.start_date))
        self.main_window.planning_tab.end_date.setDate(QDate(planning.end_date))
        
        # Mettre à jour les dates dans DesiderataManagementWidget
        self.main_window.desiderata_tab.sync_dates_from_planning(planning.start_date, planning.end_date)
        
        # Mettre à jour toutes les vues
        self.main_window.planning_tab.update_table()
        self.main_window.update_data()
        self.main_window.comparison_view.update_comparison()
        self.main_window.doctor_planning_view.update_view(planning, self.main_window.doctors, self.main_window.cats)
        self.main_window.stats_tab.update_stats(planning, self.main_window.doctors, self.main_window.cats)
        QMessageBox.information(self, "Succès", f"Planning chargé depuis {filename}")

    def export_to_csv(self):
        if not self.main_window.planning_tab.planning:
            QMessageBox.warning(self, "Erreur", "Aucun planning à exporter.")
            return

        folder = QFileDialog.getExistingDirectory(self, "Sélectionner le dossier d'exportation")
        if not folder:
            return

        planning = self.main_window.planning_tab.planning
        
        # Exporter le planning global
        self.export_global_planning_to_csv(planning, folder)

        # Exporter les plannings individuels
        for person in self.main_window.doctors + self.main_window.cats:
            self.export_individual_planning_to_csv(planning, person, folder)

        QMessageBox.information(self, "Succès", f"Plannings exportés en CSV dans {folder}")

    def export_to_excel(self):
        if not self.main_window.planning_tab.planning:
            QMessageBox.warning(self, "Erreur", "Aucun planning à exporter.")
            return

        folder = QFileDialog.getExistingDirectory(self, "Sélectionner le dossier d'exportation")
        if not folder:
            return

        planning = self.main_window.planning_tab.planning
        
        # Exporter tous les plannings dans un seul fichier Excel
        self.export_all_plannings_to_excel(planning, folder)

        QMessageBox.information(self, "Succès", f"Plannings exportés en Excel dans {folder}")

    def update_planning_list(self):
        self.planning_list.clear()
        plannings = [f for f in os.listdir() if f.startswith("P ") and f.endswith(".json")]
        self.planning_list.addItems(plannings)

    def export_global_planning_to_csv(self, planning, folder):
        filename = os.path.join(folder, f"Planning_global_{planning.start_date.strftime('%d-%m-%y')}-{planning.end_date.strftime('%d-%m-%y')}.csv")
        
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            for day in planning.days:
                date_str = day.date.strftime('%d/%m/%Y')
                slots_by_type = {}
                
                # Regrouper les slots par type de poste
                for slot in day.slots:
                    base_type = POST_TYPE_MAPPING.get(slot.abbreviation, slot.abbreviation)
                    if base_type not in slots_by_type:
                        slots_by_type[base_type] = []
                    slots_by_type[base_type].append(slot)
                
                # Écrire les lignes pour chaque type de poste ordonné
                for post_type in ORDERED_POST_TYPES:
                    base_type = post_type[:-1] if post_type[-1].isdigit() else post_type
                    index = int(post_type[-1]) if post_type[-1].isdigit() else 1
                    
                    if base_type in slots_by_type and index <= len(slots_by_type[base_type]):
                        slot = slots_by_type[base_type][index - 1]
                        writer.writerow(["+", date_str, slot.assignee or "", post_type])
                    else:
                        writer.writerow(["+", date_str, "", post_type])

    def export_individual_planning_to_csv(self, planning, person, folder):
        filename = os.path.join(folder, f"Planning_{person.name}_{planning.start_date.strftime('%d-%m-%y')}-{planning.end_date.strftime('%d-%m-%y')}.csv")
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Date', 'Créneau', 'Type', 'Site'])
            for day in planning.days:
                for slot in day.slots:
                    if slot.assignee == person.name:
                        writer.writerow([
                            day.date.strftime('%d-%m-%y'),
                            f"{slot.start_time.strftime('%H:%M')} - {slot.end_time.strftime('%H:%M')}",
                            slot.abbreviation,
                            slot.site
                        ])

    def export_all_plannings_to_excel(self, planning, folder):
        filename = os.path.join(folder, f"Tous_les_plannings_{planning.start_date.strftime('%d-%m-%y')}-{planning.end_date.strftime('%d-%m-%y')}.xlsx")
        workbook = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        workbook.remove(workbook.active)
        
        # Créer une feuille pour chaque médecin et CAT
        for person in self.main_window.doctors + self.main_window.cats:
            sheet = workbook.create_sheet(title=person.name)
            self.create_individual_planning_sheet(sheet, planning, person)

        workbook.save(filename)

    def create_individual_planning_sheet(self, sheet, planning, person):
        cal = self.cal  # Utiliser le calendrier de l'instance
    
        
        # Define colors and styles
        weekend_color = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        desiderata_color = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        weekend_desiderata_color = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Set column widths
        sheet.column_dimensions['A'].width = 5  # Jour (Numeric day column)
        sheet.column_dimensions['B'].width = 5  # Semaine column (used as a separator)

        # Create the header with months and days
        months = self.get_month_names(planning.start_date, planning.end_date)
        header = ["Jour", "Sem"]  # Add "Sem" as a standalone column for reference
        for month in months:
            header.extend([month, "", "", ""])  # Four columns for J, M, AM, S
        sheet.append(header)
        
        # Merge cells for the month headers and add subheaders for J, M, AM, S
        col = 3  # Start from column 3 to skip "Jour" and "Sem"
        for month in months:
            sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+3)
            for i, period in enumerate(["J", "M", "AM", "S"]):
                sheet.cell(row=2, column=col+i, value=period)
            col += 4  # Move to the next group of columns for the next month
        
        # Fill in the days and data
        current_date = planning.start_date
        while current_date <= planning.end_date:
            row = current_date.day
            month_col = (current_date.year - planning.start_date.year) * 12 + current_date.month - planning.start_date.month
            col = 3 + month_col * 4

            # Fill day number
            sheet.cell(row=row+2, column=1, value=current_date.day)

            # Fill weekday
            weekday = current_date.strftime("%a")[:2]
            sheet.cell(row=row+2, column=col, value=weekday)

            # Fill slots
            day_planning = next((d for d in planning.days if d.date == current_date), None)
            m, am, s = self.get_cell_values(day_planning, person)
            sheet.cell(row=row+2, column=col+1, value=m)
            sheet.cell(row=row+2, column=col+2, value=am)
            sheet.cell(row=row+2, column=col+3, value=s)

            # Apply styles and colors
            for i in range(4):
                cell = sheet.cell(row=row+2, column=col+i)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                is_weekend = current_date.weekday() >= 5
                is_holiday_or_bridge = cal.is_holiday(current_date) or self.is_bridge_day(current_date)
                period = ["J", "M", "AM", "S"][i]
                has_desiderata = self.has_desiderata(person, current_date, period)

                if is_weekend or is_holiday_or_bridge:
                    if has_desiderata:
                        cell.fill = weekend_desiderata_color
                    else:
                        cell.fill = weekend_color
                elif has_desiderata:
                    cell.fill = desiderata_color

            current_date += timedelta(days=1)

        # Set column widths for all J, M, AM, S columns (except "Sem")
        for col in range(3, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 6

        # Apply styles to the headers
        for cell in sheet[1] + sheet[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def is_bridge_day(self, date):
        # 1) Lundi avant un mardi férié
        if date.weekday() == 0 and self.cal.is_holiday(date + timedelta(days=1)):
            return True
        
        # 2) Vendredi et samedi après un jeudi férié
        if date.weekday() in [4, 5] and self.cal.is_holiday(date - timedelta(days=1 if date.weekday() == 4 else 2)):
            return True
        
        # 3) Samedi après un vendredi férié
        if date.weekday() == 5 and self.cal.is_holiday(date - timedelta(days=1)):
            return True
        
        # 4) Jour de semaine entre deux jours fériés
        if 0 <= date.weekday() <= 4:  # Jours de semaine (lundi à vendredi)
            if (self.cal.is_holiday(date - timedelta(days=1)) and 
                self.cal.is_holiday(date + timedelta(days=1))):
                return True
        
        return False
            
    def get_cell_values(self, day_planning, person):
        if not day_planning:
            return "", "", ""
        slots = [slot for slot in day_planning.slots if slot.assignee == person.name]
        periods = {"M": [], "AM": [], "S": []}
        for slot in slots:
            period = self.get_post_period(slot.abbreviation)
            periods[period].append(slot.abbreviation)
        return " ".join(periods["M"]), " ".join(periods["AM"]), " ".join(periods["S"])

    def get_post_period(self, post):
        if post in ["ML", "MC", "MM", "CM", "HM", "SM", "RM"]:
            return "M"
        elif post in ["CA", "HA", "SA", "RA", "AL", "AC"]:
            return "AM"
        else:
            return "S"

    def has_desiderata(self, person, date, period):
        for des in person.desiderata:
            if des.start_date <= date <= des.end_date:
                if (period == "M" and des.period == 1) or \
                (period == "AM" and des.period == 2) or \
                (period == "S" and des.period == 3):
                    return True
        return False

    def get_month_names(self, start_date, end_date):
        months = []
        current_date = start_date.replace(day=1)
        while current_date <= end_date:
            months.append(current_date.strftime("%b"))
            current_date = (current_date.replace(day=1) + timedelta(days=32)).replace(day=1)
        return months
